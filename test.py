import io
import uuid
import pandas as pd
import streamlit as st
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from sfc_quiz.utility.utils import handle_table_input


def excel_to_dataframe_reference(cell_ref):
    col_str, row_str = '', ''
    for char in cell_ref:
        if char.isalpha():
            col_str += char
        else:
            row_str += char
    col_num = 0
    for char in col_str:
        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
    row_num = int(row_str) - 1
    return row_num, col_num - 1


st.title("SFC Quiz Tesing App")
handle_table_input()


def run_excel_formula_app(df, session_key):
    st.title("Excel Formula Generator")

    if 'selected_formula' not in st.session_state:
        st.session_state['selected_formula'] = None

    if 'formula_type' not in st.session_state:
        st.session_state['formula_type'] = None

    formula_type = st.selectbox("Select a formula type:",
                                ["SUM", "HLOOKUP", "FILTER", "SUMIF", "VLOOKUP", "MATCH", "INDEX", "AVERAGE",
                                 "INDEX-MATCH","IF","PIVOT_TABLE","COMPLEX_IF"], key=f"formula_type_{session_key}")
    column = None
    lookup_value = None
    lookup_column = None
    range_columns = None
    col_index_num = None
    range_lookup = None
    condition = None
    condition_column = None
    condition_value = None
    condition_operator = None
    Array = None
    result_column = None
    row = None
    sum_column = None
    if formula_type == "SUM" or formula_type == "AVG":
        column = st.selectbox("Select the column for the operation:", df.columns)
    elif formula_type == "SUMIF":
        range_columns = st.multiselect("Select the range of columns for SUMIF:", df.columns,
                                       default=df.columns.tolist())
        sum_column = st.selectbox("Select the column to sum:", df.columns)
        condition_column = st.selectbox("Select the column for the condition:", df.columns)
        condition = st.text_input("Enter the condition for SUMIF (e.g., '>4000'):")
    elif formula_type == "IF":
        condition_column = st.selectbox("Select the condition column:", df.columns, key=f'cond_column_{session_key}')
        condition = st.text_input("Enter condition (e.g., '>5'):", key=f'condition_{session_key}')
        true_result = st.text_input("Enter result if true (e.g., 'High'):", key=f'true_result_{session_key}')
        false_result = st.text_input("Enter result if false (e.g., 'Low'):", key=f'false_result_{session_key}')
    elif formula_type == "FILTER":
        range_columns = st.multiselect("Select the column to filter:", df.columns, default=df.columns.tolist())
        condition_column = st.selectbox("Select the column for the condition:", df.columns)
        condition_value = st.text_input("Enter the condition value (e.g., '20'):")
        condition_operator = st.selectbox("Choose the operator:", [">", "<", "=", ">=", "<="])
    elif formula_type == "HLOOKUP":
        lookup_value = st.text_input("Enter the lookup value for HLOOKUP:")
        range_columns = st.multiselect("Select the range of rows to search within:", df.columns,
                                       default=df.columns.tolist())
        col_index_num = st.number_input("Enter the row index for the result (starting from 1):", min_value=1,
                                        max_value=len(df.columns), value=1)
        range_lookup = st.radio("Choose the type of match for HLOOKUP:", ['True', 'False'], index=1)
    if formula_type == "PIVOT_TABLE":
        values = st.multiselect("Select value column(s):", df.columns)
        index = st.multiselect("Select index column(s):", df.columns)
        columns = st.multiselect("Select column(s) to pivot:", df.columns)
        aggfunc = st.selectbox("Select aggregation function:", ["sum", "mean", "count", "min", "max"])
        pivot_options = {
            "fill_value": st.number_input("Fill missing values with:", value=0),
            "margins": st.checkbox("Show grand totals"),
            "margins_name": st.text_input("Grand total name:", value="Grand Total"),
            "sort_values": st.checkbox("Sort values"),
            "sort_ascending": st.checkbox("Sort in ascending order", value=True),
        }
    elif formula_type == "VLOOKUP":
        lookup_column = st.selectbox("Select the lookup column:", df.columns)
        lookup_value = st.text_input("Enter the lookup value:")
        range_columns = st.multiselect("Select the range of columns to search within:", df.columns,
                                       default=df.columns.tolist())
        col_index_num = st.number_input("Enter the column index for the result (starting from 1):", min_value=1,
                                        max_value=len(df.columns), value=1)
        range_lookup = st.radio("Choose the type of match:", ['True', 'False'], index=1)
    elif formula_type == "MATCH":
        column = st.selectbox("Select the column for MATCH:", df.columns)
        lookup_value = st.text_input("Enter lookup value for MATCH:")
    elif formula_type == "INDEX":
        Array = st.multiselect("Select the Array:", df.columns, default=df.columns.tolist())
        column = st.selectbox("Select the column for INDEX:", Array)
        row = st.number_input("Enter the index position (row number):", min_value=1, max_value=len(df), value=1)
    elif formula_type == "INDEX-MATCH":
        lookup_value = st.text_input("Enter the lookup value:")
        lookup_column = st.selectbox("Select the lookup column:", df.columns)
        result_column = st.selectbox("Select the result column:", df.columns)
        Array = st.multiselect("Select the Array:", df.columns, default=df.columns.tolist())
    if st.button(f"Generate {formula_type} Formula"):
        formula = ""
        if formula_type == "SUM":
            formula = f"=SUM({chr(ord('A') + df.columns.get_loc(column))}2:{chr(ord('A') + df.columns.get_loc(column))}{len(df) + 1})"
        elif formula_type == "SUMIF":
            if condition:
                start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
                end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
                sum_col_letter = chr(ord('A') + df.columns.get_loc(sum_column))
                condition_col_letter = chr(ord('A') + df.columns.get_loc(condition_column))
                formula = f'=SUMIF({condition_col_letter}2:{condition_col_letter}{len(df) + 1}, "{condition}", {sum_col_letter}2:{sum_col_letter}{len(df) + 1})'
            else:
                st.warning("Please enter a condition for SUMIF.")
        elif formula_type == "VLOOKUP":
            row_index = df[df[lookup_column].astype(str) == lookup_value].index[0] + 2
            lookup_value_cell = f'{chr(ord("A") + df.columns.get_loc(lookup_column))}{row_index}'
            start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
            end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
            table_range = f"{start_col}2:{end_col}{len(df) + 1}"
            formula = f"=VLOOKUP({lookup_value_cell}, {table_range}, {col_index_num}, {range_lookup})"
        elif formula_type == "MATCH":
            row_index = df[df[column].astype(str) == lookup_value].index[0] + 2
            lookup_value = f'{chr(ord("A") + df.columns.get_loc(column))}{row_index}'
            formula = f"=MATCH({lookup_value}, {chr(ord('A') + df.columns.get_loc(column))}2:{chr(ord('A') + df.columns.get_loc(column))}{len(df) + 1}, 0)"
        elif formula_type == "IF":
            formula = f"=IF({get_column_letter(df.columns.get_loc(condition_column) + 1)}1{condition}, \"{true_result}\", \"{false_result}\")"
        elif formula_type == "PIVOT_TABLE":
            pivot_table = pd.pivot_table(df, values=values, index=index, columns=columns, aggfunc=aggfunc)
            if values and index:
                pivot_table = pd.pivot_table(
                    df,
                    values=values,
                    index=index,
                    columns=columns,
                    aggfunc=aggfunc,
                    fill_value=pivot_options["fill_value"],
                    margins=pivot_options["margins"],
                    margins_name=pivot_options["margins_name"],
                )

                if pivot_options["sort_values"]:
                    sort_order = "ascending" if pivot_options["sort_ascending"] else "descending"
                    pivot_table = pivot_table.sort_values(by=values, ascending=pivot_options["sort_ascending"])

                st.subheader("Pivot Table Result")
                st.write(pivot_table)

                formatting_options = {
                    "highlight_max": st.checkbox("Highlight maximum values"),
                    "highlight_min": st.checkbox("Highlight minimum values"),
                    "italic_headers": st.checkbox("Italicize headers"),
                    "bold_headers": st.checkbox("Bold headers"),
                }

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    pivot_table.to_excel(writer, sheet_name='Pivot Table')
                    worksheet = writer.sheets['Pivot Table']

                    if formatting_options["highlight_max"]:
                        worksheet.conditional_formatting.add('B2:Z1000', rule=writer.book.add_format(
                            {'bg_color': '#FFC7CE', 'font_color': '#9C0006'}), rule_type='max')
                    if formatting_options["highlight_min"]:
                        worksheet.conditional_formatting.add('B2:Z1000', rule=writer.book.add_format(
                            {'bg_color': '#C6EFCE', 'font_color': '#006100'}), rule_type='min')
                    if formatting_options["italic_headers"]:
                        worksheet.set_row(0, cell_format=writer.book.add_format({'italic': True}))
                    if formatting_options["bold_headers"]:
                        worksheet.set_row(0, cell_format=writer.book.add_format({'bold': True}))

                output.seek(0)

                st.download_button(
                    label="Download Pivot Table",
                    data=output,
                    file_name="pivot_table.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("Please select at least one value column and one index column.")
        elif formula_type == "FILTER":
            condition_match = df[df[condition_column].astype(str) == condition_value]
            if not condition_match.empty:
                row_index = condition_match.index[0] + 2
                start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
                end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
                condition_col_letter = chr(ord('A') + df.columns.get_loc(condition_column))
                table_range = f"{start_col}2:{end_col}{len(df) + 1}"
                condition_value = f'{chr(ord("A") + df.columns.get_loc(condition_column))}{row_index}'
                condition_str = f"{condition_col_letter}2:{condition_col_letter}{len(df) + 1} {condition_operator} {condition_value}"
                formula = f"=FILTER({table_range}, {condition_str})"
            else:
                st.warning("No rows match the specified condition.")

            st.session_state['selected_formula'] = formula

        elif formula_type == "HLOOKUP":
            start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
            end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
            range_lookup_val = "TRUE" if range_lookup == "True" else "FALSE"
            formula = f"=HLOOKUP({lookup_value}, {start_col}1:{end_col}{len(df) + 1}, {col_index_num}, {range_lookup_val})"
        elif formula_type == "INDEX":
            if column in df.columns:
                row_num = int(row)
                start_col = chr(ord('A') + df.columns.get_loc(Array[0]))
                end_col = chr(ord('A') + df.columns.get_loc(Array[-1]))
                col_num = df.columns.get_loc(column) + 1
                formula = f"=INDEX({start_col}2:{end_col}{len(df) + 1}, {row_num}, {col_num})"
            else:
                st.warning("Selected column is not present in the DataFrame.")
        elif formula_type == "INDEX-MATCH":

            if lookup_value and lookup_column and result_column:
                lookup_col_letter = chr(ord('A') + df.columns.get_loc(lookup_column))
                result_col_letter = chr(ord('A') + df.columns.get_loc(result_column))
                formula = f'=INDEX({lookup_col_letter}2:{result_col_letter}{len(df) + 1}, MATCH({lookup_value}, {lookup_col_letter}2:{lookup_col_letter}{len(df) + 1}, 0))'
            else:
                st.warning("Please provide the lookup value, lookup column, and result column.")
        elif formula_type == "AVG":
            formula = f"=AVERAGE({chr(ord('A') + df.columns.get_loc(column))}2:{chr(ord('A') + df.columns.get_loc(column))}{len(df) + 1})"

        # elif formula_type == "COMPLEX_IF":
        #     st.subheader("Complex IF Function Configuration")
        #
        #     condition_cells = []
        #     value_inputs = []
        #
        #     num_conditions = st.number_input("Enter the number of conditions:", min_value=1, value=3, step=1)
        #
        #     for i in range(num_conditions):
        #         condition_cell = st.selectbox(f"Select condition cell {i + 1}:", df.columns, key=f"condition_cell_{i}")
        #         condition_cells.append(condition_cell)
        #
        #         value_input = st.text_input(f"Enter value {i + 1}:", key=f"value_input_{i}")
        #         value_inputs.append(value_input)

            # if st.button("Generate Complex IF Function"):
            #     complex_if_function = "=IF("
            #     for i in range(num_conditions):
            #         value_cell = xl_rowcol_to_cell(1, df.columns.get_loc(condition_cells[i]) + 1)
            #         complex_if_function += f"{condition_cells[i]}<=A1,{value_cell}"
            #         if i < num_conditions - 1:
            #             complex_if_function += ",("
            #         else:
            #             complex_if_function += ")" * (num_conditions - 1)
            #
            #     st.subheader("Generated Complex IF Function")
            #     st.code(complex_if_function)
            #
            #     if st.button("Apply Complex IF Function"):
            #         try:
            #             output = io.BytesIO()
            #             with pd.ExcelWriter(output, engine='openpyxl') as writer:
            #                 df.to_excel(writer, index=False)
            #                 workbook = writer.book
            #                 sheet = workbook.active
            #                 max_col = df.shape[1]
            #                 formula_col = get_column_letter(max_col + 3)
            #                 formula_cell = f"{formula_col}1"
            #                 sheet[formula_cell] = complex_if_function
            #
            #                 # Write the input values to the corresponding cells
            #                 for i in range(num_conditions):
            #                     value_cell = xl_rowcol_to_cell(1, df.columns.get_loc(condition_cells[i]) + 1)
            #                     sheet[value_cell] = value_inputs[i]
            #
            #                 sheet[formula_cell].font = Font(bold=True)
            #                 sheet[formula_cell].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
            #                                                        fill_type="solid")
            #
            #             output.seek(0)
            #             wb = load_workbook(output)
            #             ws = wb.active
            #             ws[formula_cell].data_type = 'f'
            #             wb.calc_mode = 'auto'
            #             result = ws[formula_cell].calculate()
            #             st.success(
            #                 f"The complex IF function has been applied at cell {formula_cell} and the result is {result}")
            #             st.download_button(
            #                 label="Download Excel file",
            #                 data=output,
            #                 file_name="modified_data.xlsx",
            #                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            #             )
            #         except Exception as e:
            #             st.error(f"Error applying the complex IF function: {str(e)}")


        st.session_state['selected_formula'] = formula
        st.session_state['formula_type'] = formula_type

    if st.session_state['selected_formula']:
        st.text("Generated Excel Formula:")
        st.code(st.session_state['selected_formula'])

        if st.button("Apply Formula"):
            formula = st.session_state['selected_formula']
            try:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                    workbook = writer.book
                    sheet = workbook.active
                    max_col = df.shape[1]  # Number of columns in the DataFrame
                    formula_col = get_column_letter(max_col + 3)  # 2 columns space + 1 for the next available column
                    formula_cell = f"{formula_col}1"
                    sheet[formula_cell] = formula
                    sheet[formula_cell].font = Font(bold=True)
                    sheet[formula_cell].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                output.seek(0)
                wb = load_workbook(output)
                ws = wb.active
                ws[formula_cell].data_type = 'f'
                wb.calc_mode = 'auto'
                result = ws[formula_cell].calculate()
                st.success(f"The formula has been applied at cell {formula_cell} and result is {result}")
                st.download_button(
                    label="Download Excel file",
                    data=output,
                    file_name="modified_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"Error applying the formula: {str(e)}")


if st.session_state['df'] is not None:
    df = st.session_state['df']
    run_excel_formula_app(df, session_key="hemanth")
