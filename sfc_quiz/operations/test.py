import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import json


def load_formulas():
    with open("excel-formulas.json", 'r') as file:
        return json.load(file)


def formula_interface(df):
    formulas = load_formulas()

    formula_names = [f['name'] for f in formulas['formulas']]
    selected_formula_name = st.selectbox("Select a formula", formula_names)
    selected_formula = next((f for f in formulas['formulas'] if f['name'] == selected_formula_name), None)
    st.write("Description:", selected_formula['description'])
    st.write(f"**Formula:** {selected_formula['formula']}")

    input_values = {}
    for input_name, input_description in selected_formula['inputs'].items():
        if input_name == 'table_array':
            start_col, end_col = st.columns(2)
            start_column = start_col.selectbox(f"Select start column for {input_description}", list(df.columns))
            end_column = end_col.selectbox(f"Select end column for {input_description}", list(df.columns))
            start_col_index = df.columns.get_loc(start_column)
            end_col_index = df.columns.get_loc(end_column)
            input_value = df.iloc[:, start_col_index:(end_col_index + 1)]
        elif input_name == 'col_index_num':
            max_col_index = len(df.columns)
            input_value = st.number_input(input_description, min_value=1, max_value=max_col_index, step=1)
        elif input_name == 'range_lookup':
            input_value = st.selectbox(input_description, ['True', 'False'])
        else:
            input_value = st.text_input(input_description)
        input_values[input_name] = input_value
    if 'range_lookup' not in input_values:
        input_values['range_lookup'] = 'True'

    if st.button("Apply Formula"):
        try:
            formula = selected_formula['formula'].format(**input_values)
            st.write(f"**Applied Formula:** {formula}")
            if formula:
                result, applied_formula = apply_formula_to_excel(df, formula, input_values)
                st.write(f"**Result in Excel:** {result}")
                st.write(f"**Applied Formula in Excel:** {applied_formula}")
        except KeyError as e:
            missing_input = str(e).strip("'")
            st.error(f"The input '{missing_input}' is missing in the formula. Please check the formula definition.")

def apply_formula_to_excel(df, formula, input_values):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    num_rows = df.shape[0]
    if 'table_array' in input_values:
        table_array = input_values['table_array']
        start_col_index = table_array.columns.get_loc(table_array.columns[0])
        end_col_index = table_array.columns.get_loc(table_array.columns[-1])
        start_col_letter = get_column_letter(start_col_index + 1)
        end_col_letter = get_column_letter(end_col_index + 1)
        table_range = f"{start_col_letter}2:{end_col_letter}{num_rows + 1}"
        formula = formula.replace('table_array', table_range)

    if 'lookup_value' in formula:
        lookup_value = input_values.get('lookup_value')
        formula = formula.replace('lookup_value', str(lookup_value))

    if 'col_index_num' in formula:
        col_index_num = input_values.get('col_index_num')
        col_index_num_letter = get_column_letter(col_index_num)
        formula = formula.replace('col_index_num', col_index_num_letter)

    result_column = ws.max_column + 1
    result_column_letter = get_column_letter(result_column)
    for row in range(2, num_rows + 2):
        cell = ws[f"{result_column_letter}{row}"]
        cell.value = f'={formula}'

    result_values = [cell[0].value for cell in ws[f"{result_column_letter}2:{result_column_letter}{num_rows + 1}"]]
    applied_formula = ws[f"{result_column_letter}2"].value


    return result_values, applied_formula