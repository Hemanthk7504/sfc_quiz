# import streamlit as st
# import matplotlib.pyplot as plt
# import plotly.express as px
#
# from formulas.excel_formulas import *
#
#
# def data_operations_module(df):
#     """Module for performing data operations (formulas and statistics) on DataFrame."""
#     if df is not None:
#         formula_choice = st.selectbox(
#             "Select a formula or operation",
#             ("VLOOKUP", "SUM", "AVERAGE", "STANDARD DEVIATION", "VARIANCE", "MODE"),
#             key=f'formula_choice_1'
#         )
#
#         if formula_choice == "VLOOKUP":
#             lookup_value = st.text_input("Enter the value to lookup:", key=f"lookup_value_{formula_choice}")
#             lookup_column = st.selectbox("Choose the column to lookup from:", df.columns, key=f"lookup_column_{formula_choice}")
#             result_column = st.selectbox("Choose the column to return value from:", df.columns, key=f"result_column_{formula_choice}")
#             if st.button("Perform Lookup", key=f"perform_lookup_{formula_choice}"):
#                 result, match_index, match_details = perform_vlookup(df, lookup_column, lookup_value, result_column)
#                 if result:
#                     st.success(f"Found value: {result} at index {match_index}")
#                     st.write("Details:", match_details)
#                 else:
#                     st.error("No matching entry found.")
#
#         elif formula_choice == "SUM":
#             selected_columns = st.multiselect("Select columns to sum (hold Ctrl for multiple):", df.columns, key=f"sum_columns_{formula_choice}")
#             if st.button("Calculate SUM", key=f"calculate_sum_{formula_choice}"):
#                 if not selected_columns:
#                     st.error("Please select at least one column.")
#                 else:
#                     results = {col: df[col].sum() for col in selected_columns}
#                     st.write("Sum of selected columns:")
#                     for col, value in results.items():
#                         st.write(f"- {col}: {value}")
#
#         elif formula_choice == "AVERAGE":
#             column_name = st.selectbox("Select the column for average calculation:", df.columns, key=f"average_column_{formula_choice}")
#             if st.button("Calculate AVERAGE", key=f"calculate_average_{formula_choice}"):
#                 result = calculate_mean(df, column_name)
#                 st.write("Average of", column_name, "is", result)
#
#         elif formula_choice == "STANDARD DEVIATION":
#             column_name = st.selectbox("Select the column for standard deviation calculation:", df.columns, key=f"std_column_{formula_choice}")
#             if st.button("Calculate STANDARD DEVIATION", key=f"calculate_std_{formula_choice}"):
#                 result = calculate_std(df, column_name)
#                 st.write(f"Standard Deviation of {column_name}: {result}")
#
#         elif formula_choice == "VARIANCE":
#             column_name = st.selectbox("Select the column for variance calculation:", df.columns, key=f"variance_column_{formula_choice}")
#             if st.button("Calculate VARIANCE", key=f"calculate_variance_{formula_choice}"):
#                 result = calculate_variance(df, column_name)
#                 st.write(f"Variance of {column_name}: {result}")
#
#         elif formula_choice == "MODE":
#             column_name = st.selectbox("Select the column for mode calculation:", df.columns, key=f"mode_column_{formula_choice}")
#             if st.button("Calculate MODE", key=f"calculate_mode_{formula_choice}"):
#                 result = calculate_mode(df, column_name)
#                 if isinstance(result, list):
#                     result = ', '.join(map(str, result))
#                 st.write(f"Mode of {column_name}: {result}")
#
#
#
# def plot_data(df):
#     graph_type = st.selectbox(
#         "Select the type of graph",
#         ("Line Chart", "Bar Chart", "Histogram", "Box Plot", "Scatter Plot", "Pie Chart")
#     )
#
#     if graph_type != "Pie Chart":
#         selected_columns = st.multiselect("Select columns to plot", df.columns)
#     else:
#         selected_columns = st.selectbox("Select one column for pie chart", df.columns)
#
#     if not selected_columns:
#         st.warning("Please select at least one column.")
#         return
#
#     if st.button("Generate Graph"):
#         if graph_type == "Line Chart":
#             plot_line_chart(df, selected_columns)
#         elif graph_type == "Bar Chart":
#             plot_bar_chart(df, selected_columns)
#         elif graph_type == "Histogram":
#             plot_histogram(df, selected_columns)
#         elif graph_type == "Box Plot":
#             plot_box_plot(df, selected_columns)
#         elif graph_type == "Scatter Plot":
#             plot_scatter_plot(df, selected_columns)
#         elif graph_type == "Pie Chart":
#             plot_pie_chart(df, selected_columns)
#
#
# def plot_line_chart(df, columns):
#     fig, ax = plt.subplots()
#     for col in columns:
#         ax.plot(df[col], label=col)
#     plt.legend()
#     st.pyplot(fig)
#
#
# def plot_bar_chart(df, columns):
#     df[columns].plot(kind='bar')
#     st.pyplot(plt)
#
#
# def plot_histogram(df, columns):
#     df[columns].hist()
#     st.pyplot(plt)
#
#
# def plot_box_plot(df, columns):
#     df[columns].plot(kind='box')
#     st.pyplot(plt)
#
#
# def plot_scatter_plot(df, columns):
#     if len(columns) < 2:
#         st.error("Please select at least two columns for scatter plot.")
#         return
#     fig = px.scatter(df, x=columns[0], y=columns[1])
#     st.plotly_chart(fig)
#
#
# def plot_pie_chart(df, column):
#     fig = px.pie(df, names=column, values=df[column])
#     st.plotly_chart(fig)


import json
import re
from io import BytesIO
from typing import Dict, Any

import pandas as pd
import streamlit as st


def load_formulas():
    with open('excel-formulas.json', 'r') as file:
        formulas = json.load(file)
    return formulas


def convert_to_cell_reference(df, col_name, row_index):
    if df is None:
        return col_name

    try:
        col_index = df.columns.get_loc(col_name)
        col_letter = chr(ord('A') + col_index)
        return f"{col_letter}{row_index + 1}"
    except KeyError:
        return col_name


def apply_formula(dataframe, formula, variables):
    try:
        for var_name, var_value in variables.items():
            if var_name in formula:
                if 'range' in var_name or 'table_array' in var_name:
                    start_col, end_col = var_value.split(':')
                    start_row = 1
                    end_row = dataframe.shape[0]
                    start_cell = convert_to_cell_reference(dataframe, start_col, start_row - 1)
                    end_cell = convert_to_cell_reference(dataframe, end_col, end_row - 1)
                    cell_range = f"{start_cell}:{end_cell}"
                    formula = formula.replace(f"{{{var_name}}}", cell_range)
                elif 'lookup_value' in var_name or 'value' in var_name:
                    if var_value in dataframe.columns:
                        cell_reference = convert_to_cell_reference(dataframe, var_value, 0)
                        formula = formula.replace(f"{{{var_name}}}", str(cell_reference))
                    else:
                        formula = formula.replace(f"{{{var_name}}}", str(var_value))
                elif 'col_index_num' in var_name:
                    col_index_num = int(var_value)
                    formula = formula.replace(f"{{{var_name}}}", str(col_index_num))
                else:
                    formula = formula.replace(f"{{{var_name}}}", str(var_value))

        return formula
    except Exception as e:
        return str(e)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def evaluate_formula(formula):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '=' + formula
    result = ws['A1'].value
    return result

def formula_interface(df):
    formulas = load_formulas()
    formula_names = [formula['name'] for formula in formulas]

    st.title("Apply Excel Formulas")
    selected_formula_name = st.selectbox("Choose a formula:", formula_names)
    selected_formula = next((item for item in formulas if item['name'] == selected_formula_name), None)

    if selected_formula:
        st.write(selected_formula['description'])
        variables: Dict[str, Any] = {}
        inputs = selected_formula.get('inputs', {})
        formula_str = selected_formula['formula']
        variable_names = re.findall(r"\{([^}]+)}", formula_str)

        for var_name in variable_names:
            input_prompt = inputs.get(var_name, f"Enter {var_name}:")
            if 'range' in var_name or 'table_array' in var_name:
                start_col = st.selectbox(f"Select the start column for {var_name}:", df.columns,
                                         key=f"start_{var_name}")
                end_col = st.selectbox(f"Select the end column for {var_name}:", df.columns, key=f"end_{var_name}")
                variables[var_name] = f"{start_col}:{end_col}"
            elif 'lookup_value' in var_name:
                lookup_col = st.selectbox(f"Select the column to look up for {var_name}:", df.columns, key=var_name)
                lookup_value = st.text_input(f"Enter the value to look up for {var_name}:", key=f"value_{var_name}")
                variables[var_name] = lookup_value
            elif 'col_index_num' in var_name:
                col_index_num = st.number_input(input_prompt, min_value=1, max_value=len(df.columns), step=1,
                                                key=var_name)
                variables[var_name] = col_index_num
            else:
                variables[var_name] = st.text_input(input_prompt, key=var_name)

        if st.button("Apply"):
            try:
                formula_with_cells = apply_formula(df, formula_str, variables)
                st.success(f"Formula: {formula_with_cells}")

                result = evaluate_formula(formula_with_cells)
                st.success(f"Result: {result}")

            except Exception as e:
                st.error(f"Error applying formula: {str(e)}")


def is_valid_cell_range(cell_range):
    pattern = r'^[A-Z]+\d+:[A-Z]+\d+$'
    return bool(re.match(pattern, cell_range))


def apply_formula(dataframe, formula, variables):
    # Secure the environment for eval
    allowed_names = {"dataframe": dataframe}
    allowed_names.update(variables)

    try:
        # Only allow the evaluation of expressions that use the dataframe and provided variables
        result = eval(formula, {"__builtins__": None}, allowed_names)
        return result
    except Exception as e:
        return str(e)