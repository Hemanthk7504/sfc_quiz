import io
import uuid
import xlwings as xw
import pandas as pd
import streamlit as st
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from sfc_quiz.utility.utils import handle_table_input


def excel_to_dataframe_reference(cell_ref):
    col_str, row_str = '', ''
    for char in cell_ref:
        if char.isalpha():
            col_str += char
        else:
            row_str += char
    col_num = 0
    for char in col_str:
        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
    row_num = int(row_str) - 1
    return row_num, col_num - 1

st.title("SFC Quiz Tesing App")

handle_table_input()


def run_excel_formula_app(df, session_key):
    st.title("Excel Formula Generator")

    if 'selected_formula' not in st.session_state:
        st.session_state['selected_formula'] = None

    if 'formula_type' not in st.session_state:
        st.session_state['formula_type'] = None

    formula_type = st.selectbox("Select a formula type:",
                                ["SUM", "HLOOKUP", "FILTER", "SUMIF", "VLOOKUP", "MATCH", "INDEX", "AVG",
                                 "INDEX-MATCH"], key=f"formula_type_{session_key}")
    column = None
    lookup_value = None
    lookup_column = None
    range_columns = None
    col_index_num = None
    range_lookup = None
    condition = None
    condition_column = None
    condition_value = None
    condition_operator = None
    Array = None
    result_column = None
    row = None
    sum_column = None
    if formula_type == "SUM" or formula_type == "AVG":
        column = st.selectbox("Select the column for the operation:", df.columns)
    elif formula_type == "SUMIF":
        range_columns = st.multiselect("Select the range of columns for SUMIF:", df.columns,
                                       default=df.columns.tolist())
        sum_column = st.selectbox("Select the column to sum:", df.columns)
        condition_column = st.selectbox("Select the column for the condition:", df.columns)
        condition = st.text_input("Enter the condition for SUMIF (e.g., '>4000'):")
    elif formula_type == "FILTER":
        range_columns = st.multiselect("Select the column to filter:", df.columns, default=df.columns.tolist())
        condition_column = st.selectbox("Select the column for the condition:", df.columns)
        condition_value = st.text_input("Enter the condition value (e.g., '20'):")
        condition_operator = st.selectbox("Choose the operator:", [">", "<", "=", ">=", "<="])
    elif formula_type == "HLOOKUP":
        lookup_value = st.text_input("Enter the lookup value for HLOOKUP:")
        range_columns = st.multiselect("Select the range of rows to search within:", df.columns,
                                       default=df.columns.tolist())
        col_index_num = st.number_input("Enter the row index for the result (starting from 1):", min_value=1,
                                        max_value=len(df.columns), value=1)
        range_lookup = st.radio("Choose the type of match for HLOOKUP:", ['True', 'False'], index=1)

    elif formula_type == "VLOOKUP":
        lookup_column = st.selectbox("Select the lookup column:", df.columns)
        lookup_value = st.text_input("Enter the lookup value:")
        range_columns = st.multiselect("Select the range of columns to search within:", df.columns,
                                       default=df.columns.tolist())
        col_index_num = st.number_input("Enter the column index for the result (starting from 1):", min_value=1,
                                        max_value=len(df.columns), value=1)
        range_lookup = st.radio("Choose the type of match:", ['True', 'False'], index=1)
    elif formula_type == "MATCH":
        column = st.selectbox("Select the column for MATCH:", df.columns)
        lookup_value = st.text_input("Enter lookup value for MATCH:")
    elif formula_type == "INDEX":
        Array = st.multiselect("Select the Array:", df.columns, default=df.columns.tolist())
        column = st.selectbox("Select the column for INDEX:", Array)
        row = st.number_input("Enter the index position (row number):", min_value=1, max_value=len(df), value=1)
    elif formula_type == "INDEX-MATCH":
        lookup_value = st.text_input("Enter the lookup value:")
        lookup_column = st.selectbox("Select the lookup column:", df.columns)
        result_column = st.selectbox("Select the result column:", df.columns)
        Array = st.multiselect("Select the Array:", df.columns, default=df.columns.tolist())
    if st.button(f"Generate {formula_type} Formula"):
        formula = ""
        if formula_type == "SUM":
            formula = f"=SUM({chr(ord('A') + df.columns.get_loc(column))}2:{chr(ord('A') + df.columns.get_loc(column))}{len(df) + 1})"
        elif formula_type == "SUMIF":
            if condition:
                start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
                end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
                sum_col_letter = chr(ord('A') + df.columns.get_loc(sum_column))
                condition_col_letter = chr(ord('A') + df.columns.get_loc(condition_column))
                formula = f'=SUMIF({condition_col_letter}2:{condition_col_letter}{len(df) + 1}, "{condition}", {sum_col_letter}2:{sum_col_letter}{len(df) + 1})'
            else:
                st.warning("Please enter a condition for SUMIF.")
        elif formula_type == "VLOOKUP":
            row_index = df[df[lookup_column].astype(str) == lookup_value].index[0] + 2
            lookup_value_cell = f'{chr(ord("A") + df.columns.get_loc(lookup_column))}{row_index}'
            start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
            end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
            table_range = f"{start_col}2:{end_col}{len(df) + 1}"
            formula = f"=VLOOKUP({lookup_value_cell}, {table_range}, {col_index_num}, {range_lookup})"
        elif formula_type == "MATCH":
            row_index = df[df[column].astype(str) == lookup_value].index[0] + 2
            lookup_value = f'{chr(ord("A") + df.columns.get_loc(column))}{row_index}'
            formula = f"=MATCH({lookup_value}, {chr(ord('A') + df.columns.get_loc(column))}2:{chr(ord('A') + df.columns.get_loc(column))}{len(df) + 1}, 0)"
        elif formula_type == "FILTER":
            condition_match = df[df[condition_column].astype(str) == condition_value]
            if not condition_match.empty:
                row_index = condition_match.index[0] + 2
                start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
                end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
                condition_col_letter = chr(ord('A') + df.columns.get_loc(condition_column))
                table_range = f"{start_col}2:{end_col}{len(df) + 1}"
                condition_value = f'{chr(ord("A") + df.columns.get_loc(condition_column))}{row_index}'
                condition_str = f"{condition_col_letter}2:{condition_col_letter}{len(df) + 1} {condition_operator} {condition_value}"
                formula = f"=FILTER({table_range}, {condition_str})"
            else:
                st.warning("No rows match the specified condition.")

            st.session_state['selected_formula'] = formula
        elif formula_type == "HLOOKUP":
            start_col = chr(ord('A') + df.columns.get_loc(range_columns[0]))
            end_col = chr(ord('A') + df.columns.get_loc(range_columns[-1]))
            range_lookup_val = "TRUE" if range_lookup == "True" else "FALSE"
            formula = f"=HLOOKUP({lookup_value}, {start_col}1:{end_col}{len(df) + 1}, {col_index_num}, {range_lookup_val})"
        elif formula_type == "INDEX":
            if column in df.columns:
                row_num = int(row)
                start_col = chr(ord('A') + df.columns.get_loc(Array[0]))
                end_col = chr(ord('A') + df.columns.get_loc(Array[-1]))
                col_num = df.columns.get_loc(column) + 1
                formula = f"=INDEX({start_col}2:{end_col}{len(df) + 1}, {row_num}, {col_num})"
            else:
                st.warning("Selected column is not present in the DataFrame.")
        elif formula_type == "INDEX-MATCH":

            if lookup_value and lookup_column and result_column:
                lookup_col_letter = chr(ord('A') + df.columns.get_loc(lookup_column))
                result_col_letter = chr(ord('A') + df.columns.get_loc(result_column))
                formula = f'=INDEX({lookup_col_letter}2:{result_col_letter}{len(df) + 1}, MATCH({lookup_value}, {lookup_col_letter}2:{lookup_col_letter}{len(df) + 1}, 0))'
            else:
                st.warning("Please provide the lookup value, lookup column, and result column.")
        elif formula_type == "AVG":
            formula = f"=AVERAGE({chr(ord('A') + df.columns.get_loc(column))}2:{chr(ord('A') + df.columns.get_loc(column))}{len(df) + 1})"

        st.session_state['selected_formula'] = formula
        st.session_state['formula_type'] = formula_type

    if st.session_state['selected_formula']:
        st.text("Generated Excel Formula:")
        st.code(st.session_state['selected_formula'])

        if st.button("Apply Formula"):
            formula = st.session_state['selected_formula']
            try:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                    workbook = writer.book
                    sheet = workbook.active
                    max_col = df.shape[1]  # Number of columns in the DataFrame
                    formula_col = get_column_letter(max_col + 3)  # 2 columns space + 1 for the next available column
                    formula_cell = f"{formula_col}1"
                    sheet[formula_cell] = formula
                    sheet[formula_cell].font = Font(bold=True)
                    sheet[formula_cell].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                output.seek(0)
                wb = load_workbook(output)
                ws = wb.active
                result = ws[formula_cell].value
                st.success(f"The formula has been applied at cell {formula_cell}")
                st.download_button(
                    label="Download Excel file",
                    data=output,
                    file_name="modified_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"Error applying the formula: {str(e)}")


if st.session_state['df'] is not None:
    df = st.session_state['df']
    run_excel_formula_app(df, session_key="Xqwedfghjkl")
